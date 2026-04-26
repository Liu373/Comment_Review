

"""explore_comments.py
======================
Run this BEFORE override_reason_categorizer.py to inspect your dataset and:

  1. See the most frequent words — so you can decide which to add to STOPWORDS
  2. See all detected abbreviations with example sentences — so you can
     define their meanings in the ABBREVIATIONS dict in config.py

Usage:
    python explore_comments.py

Output:
  - Printed report in the console
  - explore_comments_report.xlsx  (same content, easier to review)
"""

from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

import numpy as np
import pandas as pd

import config
from override_reason_categorizer import get_column, load_sheet


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _load_reasons(file_path: Path) -> pd.Series:
    df = load_sheet(file_path, config.SHEET_NAME)
    reasons = get_column(df, "reason_text").astype(str).str.strip()
    flags   = get_column(df, "override_flag").astype(str).str.strip().str.upper()
    mask = reasons.str.len() >= config.MIN_REASON_LENGTH
    if config.ONLY_OVERRIDE_ROWS:
        mask &= flags.isin(config.OVERRIDE_YES_VALUES)
    return reasons[mask].reset_index(drop=True)


def _tokenize(text: str) -> list[str]:
    words = re.findall(r"[a-zA-Z][a-zA-Z\-']+", text.lower())
    return [w for w in words if w not in config.STOPWORDS and len(w) > 2]


def _tokenize_all(reasons: pd.Series) -> list[str]:
    tokens = []
    for text in reasons:
        tokens.extend(_tokenize(text))
    return tokens


def _detect_abbreviations(reasons: pd.Series) -> dict[str, list[str]]:
    abbr_examples: dict[str, list[str]] = {}
    pattern = re.compile(r"\b([A-Z]{2,6})\b")
    for text in reasons:
        for match in pattern.finditer(str(text)):
            abbr = match.group(1)
            if abbr not in abbr_examples:
                abbr_examples[abbr] = []
            if len(abbr_examples[abbr]) < 3:
                abbr_examples[abbr].append(text.strip())
    return abbr_examples


# ---------------------------------------------------------------------------
# Technique A — Stem groups
# Strips common suffixes to find morphological variants of the same root.
# e.g.  restructure / restructuring / restructured  →  all share root "restructur"
# ---------------------------------------------------------------------------

_SUFFIXES = ["ings", "ing", "tion", "tions", "ment", "ments",
             "ers", "ies", "ed", "es", "er", "ly", "al", "s"]

def _stem(word: str) -> str:
    for suf in _SUFFIXES:
        if word.endswith(suf) and len(word) - len(suf) >= 4:
            return word[: -len(suf)]
    return word


def _stem_groups(tokens: list[str], min_variants: int = 2) -> pd.DataFrame:
    """
    Group words that share the same stem.
    Only show groups with at least min_variants distinct surface forms.
    """
    freq  = Counter(tokens)
    stems: dict[str, set[str]] = {}
    for word in freq:
        s = _stem(word)
        stems.setdefault(s, set()).add(word)

    rows = []
    for stem, variants in stems.items():
        if len(variants) < min_variants:
            continue
        # Sort variants by frequency descending
        sorted_v = sorted(variants, key=lambda w: -freq[w])
        already  = any(
            any(v in phrases for phrases in config.SYNONYMS.values())
            for v in sorted_v
        )
        rows.append({
            "StemRoot":       stem,
            "Variants":       ", ".join(sorted_v),
            "VariantCount":   len(variants),
            "TotalFrequency": sum(freq[v] for v in variants),
            "AlreadyInSynonyms": "YES" if already else "",
            "Suggestion":     f'Add to SYNONYMS: "{stem}": [{", ".join(repr(v) for v in sorted_v)}]',
        })

    return (
        pd.DataFrame(rows)
        .sort_values("TotalFrequency", ascending=False)
        .reset_index(drop=True)
    )


# ---------------------------------------------------------------------------
# Technique B — Collocation / PMI
# Finds two-word phrases that appear together much more than by chance.
# High PMI score = strong collocation = likely a meaningful phrase unit.
# e.g.  "undrawn" + "debt" always appear together → add "undrawn debt" to SYNONYMS
# ---------------------------------------------------------------------------

def _collocations(reasons: pd.Series, top_n: int = 40, min_count: int = 2) -> pd.DataFrame:
    """
    Compute Pointwise Mutual Information (PMI) for all word bigrams.
    PMI(w1,w2) = log( P(w1,w2) / (P(w1) * P(w2)) )
    Higher PMI = words appear together more than chance.
    """
    unigram_count: Counter = Counter()
    bigram_count:  Counter = Counter()
    total_tokens = 0

    for text in reasons:
        tokens = _tokenize(text)
        total_tokens += len(tokens)
        unigram_count.update(tokens)
        for i in range(len(tokens) - 1):
            bigram_count[(tokens[i], tokens[i + 1])] += 1

    if total_tokens == 0:
        return pd.DataFrame()

    rows = []
    for (w1, w2), cnt in bigram_count.items():
        if cnt < min_count:
            continue
        p_w1 = unigram_count[w1] / total_tokens
        p_w2 = unigram_count[w2] / total_tokens
        p_pair = cnt / total_tokens
        pmi = np.log(p_pair / (p_w1 * p_w2 + 1e-12))

        phrase = f"{w1} {w2}"
        already = any(
            phrase in phrases for phrases in config.SYNONYMS.values()
        )
        rows.append({
            "Phrase":           phrase,
            "Count":            cnt,
            "PMI_Score":        round(pmi, 2),
            "Word1":            w1,
            "Word2":            w2,
            "AlreadyInSynonyms": "YES" if already else "",
            "Suggestion":       (
                f'Consider adding "{phrase}" to a SYNONYMS group'
                if pmi > 3.0 else ""
            ),
        })

    return (
        pd.DataFrame(rows)
        .sort_values("PMI_Score", ascending=False)
        .head(top_n)
        .reset_index(drop=True)
    )


# ---------------------------------------------------------------------------
# Technique C — Context similarity
# Words that appear in similar sets of comments are likely related concepts.
# Build a word × document matrix, compute cosine similarity between word vectors.
# Similar words = potential synonyms to group together.
# ---------------------------------------------------------------------------

def _context_similarity(
    reasons: pd.Series,
    top_words: int = 60,
    top_pairs: int = 40,
    min_doc_freq: int = 2,
) -> pd.DataFrame:
    """
    For each pair of frequent words, measure how similar their document
    co-occurrence patterns are (cosine similarity).
    Score near 1.0 = words appear in almost the same comments → likely synonyms.
    """
    # Build vocabulary of most frequent non-stopword words
    all_tokens = _tokenize_all(reasons)
    freq = Counter(all_tokens)

    # Document frequency
    doc_freq: dict[str, int] = {}
    for text in reasons:
        for w in set(_tokenize(text)):
            doc_freq[w] = doc_freq.get(w, 0) + 1

    vocab = [
        w for w, _ in freq.most_common(top_words * 2)
        if doc_freq.get(w, 0) >= min_doc_freq
    ][:top_words]

    if len(vocab) < 2:
        return pd.DataFrame()

    vocab_idx = {w: i for i, w in enumerate(vocab)}
    n_docs = len(reasons)

    # Word × document matrix (binary: word appears in doc?)
    W = np.zeros((len(vocab), n_docs), dtype=np.float32)
    for j, text in enumerate(reasons):
        for w in set(_tokenize(text)):
            if w in vocab_idx:
                W[vocab_idx[w], j] = 1.0

    # L2-normalise each word vector
    norms = np.linalg.norm(W, axis=1, keepdims=True)
    norms[norms == 0] = 1.0
    W /= norms

    # Cosine similarity matrix
    sim = W @ W.T   # (vocab × vocab)

    rows = []
    seen = set()
    # Iterate upper triangle, sorted by similarity
    indices = np.argwhere(sim > 0.3)
    scored  = [(float(sim[i, j]), i, j) for i, j in indices if i < j]
    scored.sort(reverse=True)

    for score, i, j in scored[:top_pairs]:
        w1, w2 = vocab[i], vocab[j]
        pair_key = tuple(sorted([w1, w2]))
        if pair_key in seen:
            continue
        seen.add(pair_key)

        already = any(
            (w1 in phrases or w2 in phrases)
            for phrases in config.SYNONYMS.values()
        )
        rows.append({
            "Word1":               w1,
            "Word2":               w2,
            "ContextSimilarity":   round(score, 3),
            "Freq_Word1":          freq[w1],
            "Freq_Word2":          freq[w2],
            "AlreadyInSynonyms":   "YES" if already else "",
            "Suggestion": (
                f'Consider grouping "{w1}" and "{w2}" in SYNONYMS'
                if score > 0.5 and not already else ""
            ),
        })

    return pd.DataFrame(rows).reset_index(drop=True)


# ---------------------------------------------------------------------------
# Original report builders
# ---------------------------------------------------------------------------

def _word_frequency_report(tokens: list[str], reasons: pd.Series) -> pd.DataFrame:
    total_docs = len(reasons)
    count = Counter(tokens)
    doc_freq: dict[str, int] = {}
    for text in reasons:
        seen = set(re.findall(r"[a-zA-Z][a-zA-Z\-']+", text.lower()))
        for w in seen:
            doc_freq[w] = doc_freq.get(w, 0) + 1

    rows = []
    for word, freq in count.most_common(80):
        if len(word) < 3:
            continue
        rows.append({
            "Word":               word,
            "TotalOccurrences":   freq,
            "InNReasons":         doc_freq.get(word, 0),
            "PctOfReasons":       round(doc_freq.get(word, 0) / total_docs * 100, 1),
            "AlreadyInStopwords": "YES" if word in config.STOPWORDS else "",
            "Suggestion": (
                "Consider adding to STOPWORDS"
                if doc_freq.get(word, 0) / total_docs > 0.4
                   and word not in config.STOPWORDS
                else ""
            ),
        })
    return pd.DataFrame(rows)


def _abbreviation_report(abbr_examples: dict[str, list[str]]) -> pd.DataFrame:
    known = {k.upper() for k in config.ABBREVIATIONS}
    rows = []
    for abbr, examples in sorted(abbr_examples.items(), key=lambda x: -len(x[1])):
        rows.append({
            "Abbreviation":   abbr,
            "ExampleCount":   len(examples),
            "AlreadyDefined": "YES" if abbr in known else "",
            "CurrentMeaning": config.ABBREVIATIONS.get(abbr, config.ABBREVIATIONS.get(abbr.upper(), "")),
            "YourMeaning":    "",
            "Example1":       examples[0] if len(examples) > 0 else "",
            "Example2":       examples[1] if len(examples) > 1 else "",
            "Example3":       examples[2] if len(examples) > 2 else "",
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Console print
# ---------------------------------------------------------------------------

def _print_section(title: str) -> None:
    print(f"\n{'='*65}")
    print(f"  {title}")
    print(f"{'='*65}")


def _print_synonym_suggestions(stem_df, colloc_df, context_df) -> None:
    _print_section("SYNONYM SUGGESTIONS — A) Stem groups (morphological variants)")
    if stem_df.empty:
        print("  No stem groups found.")
    else:
        for _, r in stem_df.head(15).iterrows():
            flag = " [already in SYNONYMS]" if r["AlreadyInSynonyms"] else ""
            print(f"  root '{r['StemRoot']}': {r['Variants']}{flag}")

    _print_section("SYNONYM SUGGESTIONS — B) Collocations (PMI — phrases that go together)")
    if colloc_df.empty:
        print("  Not enough data for collocation analysis.")
    else:
        for _, r in colloc_df[colloc_df["PMI_Score"] > 3].head(20).iterrows():
            flag = " [already in SYNONYMS]" if r["AlreadyInSynonyms"] else ""
            print(f"  '{r['Phrase']}'  (count={r['Count']}, PMI={r['PMI_Score']}){flag}")

    _print_section("SYNONYM SUGGESTIONS — C) Context similarity (words used in similar comments)")
    if context_df.empty:
        print("  Not enough data for context similarity analysis.")
    else:
        for _, r in context_df[context_df["ContextSimilarity"] > 0.5].head(20).iterrows():
            flag = " [already in SYNONYMS]" if r["AlreadyInSynonyms"] else ""
            print(f"  '{r['Word1']}' ≈ '{r['Word2']}'  (similarity={r['ContextSimilarity']}){flag}")


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def _export_report(
    word_df:    pd.DataFrame,
    abbr_df:    pd.DataFrame,
    stem_df:    pd.DataFrame,
    colloc_df:  pd.DataFrame,
    context_df: pd.DataFrame,
    output_path: Path,
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        abbr_df.to_excel(writer,    sheet_name="Abbreviations",      index=False)
        word_df.to_excel(writer,    sheet_name="Word Frequencies",   index=False)
        stem_df.to_excel(writer,    sheet_name="Synonym - Stems",    index=False)
        colloc_df.to_excel(writer,  sheet_name="Synonym - Phrases",  index=False)
        context_df.to_excel(writer, sheet_name="Synonym - Context",  index=False)

        pd.DataFrame({"Step": [
            "1 — Abbreviations sheet: fill in 'YourMeaning', add to config.py ABBREVIATIONS",
            "2 — Word Frequencies sheet: add high-frequency filler words to config.py STOPWORDS",
            "3 — Synonym - Stems sheet: word variants with same root → add all variants to one SYNONYMS group",
            "4 — Synonym - Phrases sheet: high-PMI phrases → add to SYNONYMS as multi-word entries",
            "5 — Synonym - Context sheet: similar words → consider grouping in SYNONYMS",
            "6 — Run: python override_reason_categorizer.py",
        ]}).to_excel(writer, sheet_name="Instructions", index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(output_path)
    fills = {
        "Abbreviations":     "FFF2CC",
        "Word Frequencies":  "D9EAD3",
        "Synonym - Stems":   "CFE2F3",
        "Synonym - Phrases": "FCE5CD",
        "Synonym - Context": "E8D5F5",
        "Instructions":      "D9EAF7",
    }
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        hf = fills.get(ws.title, "D9EAF7")
        for cell in ws[1]:
            cell.fill = PatternFill(fill_type="solid", fgColor=hf)
            cell.font = Font(bold=True)
        for col_cells in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col_cells
            )
            ws.column_dimensions[
                get_column_letter(col_cells[0].column)
            ].width = min(max(max_len + 2, 12), 80)
    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    base        = Path(__file__).resolve().parent
    input_path  = base / config.INPUT_FILE
    output_path = base / "explore_comments_report.xlsx"

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}\nUpdate INPUT_FILE in config.py.")

    print(f"Loading: {input_path.name}")
    reasons = _load_reasons(input_path)
    print(f"  {len(reasons):,} override reason rows found")

    if reasons.empty:
        print("No rows matched the filters. Check ONLY_OVERRIDE_ROWS / OVERRIDE_YES_VALUES in config.py.")
        return

    print("\nAnalysing...")
    tokens     = _tokenize_all(reasons)
    abbr_dict  = _detect_abbreviations(reasons)
    word_df    = _word_frequency_report(tokens, reasons)
    abbr_df    = _abbreviation_report(abbr_dict)
    stem_df    = _stem_groups(tokens)
    colloc_df  = _collocations(reasons)
    context_df = _context_similarity(reasons)

    _print_synonym_suggestions(stem_df, colloc_df, context_df)

    _export_report(word_df, abbr_df, stem_df, colloc_df, context_df, output_path)

    print(f"\n{'='*65}")
    print(f"  Report saved: {output_path}")
    print(f"{'='*65}")
    print("\nNext steps:")
    print("  1. Open explore_comments_report.xlsx")
    print("  2. Review all 5 sheets — update config.py ABBREVIATIONS, STOPWORDS, SYNONYMS")
    print("  3. Run: python override_reason_categorizer.py")


if __name__ == "__main__":
    main()




