"""Override Reason Categorizer
=================================
Automatically groups adjudicator override reasons into thematic clusters
using TF-IDF vectorisation + K-Means clustering.

Runs 100% locally — no internet connection or AI API required.

Preferred: pip install scikit-learn   (faster, more robust)
Fallback:  pure numpy/pandas implementation is included if scikit-learn is
           not available on your machine.

Usage:
    python override_reason_categorizer.py

Output: override_reason_summary_output.xlsx  (three sheets)
  - Detailed Results  : one row per override, with assigned cluster
  - Group Summary     : one row per cluster, with counts, notch stats,
                        top keywords, EntityIds, and sample reasons
  - Entity Rollup     : per-cluster breakdown by individual EntityId
"""

from __future__ import annotations

import re
import sys
from collections import Counter
from pathlib import Path
from typing import List, Optional, Tuple

import numpy as np
import pandas as pd

import config


# ---------------------------------------------------------------------------
# scikit-learn (optional — faster TF-IDF and KMeans)
# ---------------------------------------------------------------------------
try:
    from sklearn.feature_extraction.text import TfidfVectorizer as _SklearnTfidf
    from sklearn.cluster import KMeans as _SklearnKMeans
    SKLEARN_AVAILABLE = True
except ImportError:
    SKLEARN_AVAILABLE = False


# ===========================================================================
# Column helpers
# ===========================================================================

def _excel_col_to_index(col: str) -> int:
    col = col.strip().upper()
    result = 0
    for ch in col:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Invalid Excel column letter: '{col}'")
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result - 1


def _normalize_header(text: object) -> str:
    return re.sub(r"\s+", " ", str(text).strip().lower()) if text is not None else ""


def get_column(df: pd.DataFrame, key: str) -> pd.Series:
    spec = config.COLUMN_MAP[key]
    if isinstance(spec, str) and re.fullmatch(r"[A-Za-z]+", spec.strip()):
        idx = _excel_col_to_index(spec)
        if idx >= len(df.columns):
            raise KeyError(
                f"Column '{spec}' (for '{key}') is out of range — "
                f"workbook has {len(df.columns)} columns."
            )
        return df.iloc[:, idx]
    # Header-name match
    norm = {_normalize_header(c): c for c in df.columns}
    h = _normalize_header(spec)
    if h not in norm:
        raise KeyError(
            f"Header '{spec}' (for '{key}') not found. "
            f"Available headers: {list(df.columns)}"
        )
    return df[norm[h]]


# ===========================================================================
# Text cleaning
# ===========================================================================

def expand_abbreviations(text: str) -> str:
    """Replace known abbreviations with their full form (case-insensitive)."""
    for abbr, full in config.ABBREVIATIONS.items():
        text = re.sub(rf"\b{re.escape(abbr)}\b", full, text, flags=re.IGNORECASE)
    return text


# Pre-sort synonym phrases longest-first so "undrawn debt" is matched before
# just "debt". Built once at import time for efficiency.
_SYNONYM_PATTERNS: list[tuple[re.Pattern, str]] = []
for _canonical, _phrases in config.SYNONYMS.items():
    for _phrase in sorted(_phrases, key=len, reverse=True):
        _SYNONYM_PATTERNS.append(
            (re.compile(rf"\b{re.escape(_phrase)}\b", re.IGNORECASE), _canonical)
        )


def normalize_synonyms(text: str) -> str:
    """Replace all synonym phrases with their canonical key token.

    e.g. "undrawn debt" and "new loan" both become "additional_debt"
    so TF-IDF treats them as the same feature.
    """
    for pattern, canonical in _SYNONYM_PATTERNS:
        text = pattern.sub(canonical, text)
    return text


def clean_reason(text: object) -> str:
    if pd.isna(text):
        return ""
    s = re.sub(r"\s+", " ", str(text).strip())
    s = expand_abbreviations(s)
    s = normalize_synonyms(s)
    return s


def tokenize(text: str) -> List[str]:
    """Split text into lowercase tokens, removing stopwords and short words."""
    tokens = re.findall(r"[a-zA-Z][a-zA-Z\-']+", text.lower())
    return [t for t in tokens if t not in config.STOPWORDS and len(t) > 2]


# ===========================================================================
# TF-IDF (numpy fallback)
# ===========================================================================

def _build_ngrams(tokens: List[str], lo: int, hi: int) -> List[str]:
    out = []
    for n in range(lo, hi + 1):
        out.extend(" ".join(tokens[i: i + n]) for i in range(len(tokens) - n + 1))
    return out


def _tfidf_numpy(
    docs: List[str],
    stop_words: set,
    ngram_range: Tuple[int, int],
    max_features: int,
) -> Tuple[np.ndarray, List[str]]:
    """Pure numpy TF-IDF returning (X, feature_names)."""
    lo, hi = ngram_range

    def doc_terms(text: str) -> List[str]:
        tokens = [t for t in re.findall(r"[a-zA-Z][a-zA-Z\-']+", text.lower())
                  if t not in stop_words and len(t) > 2]
        return _build_ngrams(tokens, lo, hi)

    tokenized = [doc_terms(d) for d in docs]

    # Document frequency
    df_count: dict[str, int] = {}
    for terms in tokenized:
        for t in set(terms):
            df_count[t] = df_count.get(t, 0) + 1

    # Keep top max_features terms by df (then alpha for determinism)
    vocab_terms = sorted(df_count, key=lambda t: (-df_count[t], t))[:max_features]
    vocab = {t: i for i, t in enumerate(vocab_terms)}
    n_docs = len(docs)
    n_terms = len(vocab)

    X = np.zeros((n_docs, n_terms), dtype=np.float32)
    for i, terms in enumerate(tokenized):
        counts: dict[str, int] = {}
        for t in terms:
            if t in vocab:
                counts[t] = counts.get(t, 0) + 1
        total = sum(counts.values()) or 1
        for t, c in counts.items():
            tf = c / total
            idf = np.log((n_docs + 1) / (df_count[t] + 1)) + 1.0
            X[i, vocab[t]] = tf * idf

    # L2 normalise rows
    norms = np.linalg.norm(X, axis=1, keepdims=True)
    norms[norms == 0] = 1.0
    X /= norms

    return X, vocab_terms


# ===========================================================================
# K-Means (numpy fallback)
# ===========================================================================

def _kmeans_numpy(
    X: np.ndarray,
    n_clusters: int,
    n_init: int = 10,
    max_iter: int = 150,
    random_state: int = 42,
) -> Tuple[np.ndarray, np.ndarray]:
    """Return (labels, centers) using vectorised numpy K-Means."""
    rng = np.random.default_rng(random_state)
    n = X.shape[0]
    best_labels: Optional[np.ndarray] = None
    best_centers: Optional[np.ndarray] = None
    best_inertia = np.inf

    for _ in range(n_init):
        idx = rng.choice(n, n_clusters, replace=False)
        centers = X[idx].copy()

        for _ in range(max_iter):
            # Squared Euclidean distances (vectorised)
            x_sq = (X ** 2).sum(axis=1, keepdims=True)          # (n, 1)
            c_sq = (centers ** 2).sum(axis=1)                    # (k,)
            xc = X @ centers.T                                   # (n, k)
            dists_sq = np.maximum(x_sq + c_sq - 2.0 * xc, 0.0)  # (n, k)
            labels = dists_sq.argmin(axis=1)

            new_centers = np.zeros_like(centers)
            for k in range(n_clusters):
                mask = labels == k
                new_centers[k] = X[mask].mean(axis=0) if mask.any() else centers[k]

            if np.allclose(centers, new_centers, atol=1e-6):
                break
            centers = new_centers

        inertia = float(dists_sq[np.arange(n), labels].sum())
        if inertia < best_inertia:
            best_inertia = inertia
            best_labels = labels.copy()
            best_centers = centers.copy()

    return best_labels, best_centers  # type: ignore[return-value]


def _silhouette_numpy(X: np.ndarray, labels: np.ndarray) -> float:
    """Mean silhouette coefficient (vectorised, float32-safe)."""
    unique = np.unique(labels)
    if len(unique) <= 1:
        return -1.0
    n = len(X)
    scores = np.zeros(n, dtype=np.float64)
    for i in range(n):
        same = X[labels == labels[i]]
        a = np.linalg.norm(X[i] - same, axis=1).mean() if len(same) > 1 else 0.0
        b_vals = [
            np.linalg.norm(X[i] - X[labels == lbl], axis=1).mean()
            for lbl in unique if lbl != labels[i]
        ]
        b = min(b_vals)
        denom = max(a, b)
        scores[i] = (b - a) / denom if denom > 0 else 0.0
    return float(scores.mean())


# ===========================================================================
# Vectorise + Cluster
# ===========================================================================

def _vectorize(texts: List[str]) -> Tuple[np.ndarray, List[str]]:
    if SKLEARN_AVAILABLE:
        vec = _SklearnTfidf(
            ngram_range=config.NGRAM_RANGE,
            max_features=config.MAX_TFIDF_FEATURES,
            min_df=1,
            stop_words=list(config.STOPWORDS),
            sublinear_tf=True,
        )
        X = vec.fit_transform(texts).toarray().astype(np.float32)
        return X, list(vec.get_feature_names_out())
    else:
        return _tfidf_numpy(
            texts,
            config.STOPWORDS,
            config.NGRAM_RANGE,
            config.MAX_TFIDF_FEATURES,
        )


def _cluster(X: np.ndarray, n_clusters: int) -> Tuple[np.ndarray, np.ndarray]:
    if SKLEARN_AVAILABLE:
        km = _SklearnKMeans(n_clusters=n_clusters, random_state=42, n_init=10)
        labels = km.fit_predict(X)
        return labels, km.cluster_centers_
    else:
        return _kmeans_numpy(X, n_clusters)


def _choose_k(X: np.ndarray) -> int:
    """Select K by silhouette score over [MIN_K, N_CLUSTERS]."""
    best_k, best_score = config.MIN_K, -1.0
    for k in range(config.MIN_K, config.N_CLUSTERS + 1):
        if k >= X.shape[0]:
            break
        labels, _ = _cluster(X, k)
        score = _silhouette_numpy(X, labels)
        print(f"  K={k}  silhouette={score:.3f}")
        if score > best_score:
            best_score, best_k = score, k
    print(f"  → Best K = {best_k}  (score={best_score:.3f})")
    return best_k


def _cluster_label(centroid: np.ndarray, feature_names: List[str]) -> str:
    """Build a human-readable label from the top-weighted centroid features."""
    top_idx = centroid.argsort()[-config.TOP_KEYWORDS_PER_CLUSTER:][::-1]
    top_terms = [feature_names[i] for i in top_idx if centroid[i] > 0]
    return ", ".join(top_terms) if top_terms else "miscellaneous"


def cluster_reasons(texts: List[str]) -> List[str]:
    """
    Given a list of reason strings, return a cluster label for each.
    Labels are built from the most characteristic TF-IDF terms per cluster.
    """
    if not texts:
        return []

    n_docs = len(texts)
    n_clusters = min(config.N_CLUSTERS, n_docs)

    print(f"\nVectorising {n_docs} reason texts...")
    X, feature_names = _vectorize(texts)

    if config.AUTO_SELECT_K and n_docs >= config.MIN_K * 2:
        print("Auto-selecting K (silhouette)...")
        n_clusters = _choose_k(X)
    else:
        n_clusters = min(n_clusters, n_docs)

    print(f"Clustering into {n_clusters} groups...")
    labels, centers = _cluster(X, n_clusters)

    # Build a readable label for each cluster number
    cluster_label_map = {
        i: f"Group {i + 1}: {_cluster_label(centers[i], feature_names)}"
        for i in range(n_clusters)
    }

    # Print a quick cluster overview to console for sanity-checking
    print("\nCluster labels discovered:")
    for i, lbl in cluster_label_map.items():
        cnt = int((labels == i).sum())
        print(f"  [{cnt:>3}]  {lbl}")

    return [cluster_label_map[int(lbl)] for lbl in labels]


# ===========================================================================
# Load & prepare data
# ===========================================================================

def load_sheet(file_path: Path, sheet_name: Optional[str]) -> pd.DataFrame:
    kwargs: dict = {}
    if sheet_name:
        kwargs["sheet_name"] = sheet_name
    return pd.read_excel(file_path, **kwargs)


def prepare_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame()
    out["BorrowerName"]     = get_column(df, "borrower_name")
    out["EntityId"]         = get_column(df, "entity_id")
    out["ActualRating"]     = get_column(df, "actual_rating")
    out["ModelRating"]      = get_column(df, "model_rating")
    out["NotchDifference"]  = pd.to_numeric(get_column(df, "notch_difference"), errors="coerce")
    out["OverrideFlag"]     = get_column(df, "override_flag").astype(str).str.strip().str.upper()
    out["ReasonText"]       = get_column(df, "reason_text").apply(clean_reason)

    # Drop rows without a meaningful reason
    out = out[out["ReasonText"].str.len() >= config.MIN_REASON_LENGTH].copy()

    # Optionally restrict to override=Y rows
    if config.ONLY_OVERRIDE_ROWS:
        out = out[out["OverrideFlag"].isin(config.OVERRIDE_YES_VALUES)].copy()

    if out.empty:
        print("WARNING: No rows matched the filters. Check ONLY_OVERRIDE_ROWS and OVERRIDE_YES_VALUES in config.py.")
        out["Category"] = pd.Series(dtype=str)
        return out

    out["Category"] = cluster_reasons(out["ReasonText"].tolist())
    return out.reset_index(drop=True)


# ===========================================================================
# Build output tables
# ===========================================================================

def build_summary(detail: pd.DataFrame) -> pd.DataFrame:
    """One row per cluster with aggregated stats."""
    if detail.empty:
        return pd.DataFrame(columns=[
            "Category", "Count", "UniqueBorrowers",
            "AvgNotchDiff", "MedianNotchDiff", "MinNotchDiff", "MaxNotchDiff",
            "EntityIds", "TopKeywords", "SampleReasons",
        ])

    rows = []
    for category, g in detail.groupby("Category", sort=False):
        notch = g["NotchDifference"].dropna()
        entity_ids = (
            g["EntityId"].dropna().astype(str).unique().tolist()
            [: config.MAX_ENTITY_IDS_IN_SUMMARY_CELL]
        )
        samples = (
            g["ReasonText"].dropna().unique().tolist()
            [: config.MAX_SAMPLE_REASONS_PER_GROUP]
        )

        # Extract top keywords from reasons within this cluster
        c = Counter()
        for text in g["ReasonText"]:
            c.update(tokenize(text))
        top_kw = ", ".join(w for w, _ in c.most_common(8))

        rows.append({
            "Category":         category,
            "Count":            len(g),
            "UniqueBorrowers":  g["EntityId"].nunique(dropna=True),
            "AvgNotchDiff":     round(notch.mean(), 2)   if len(notch) else None,
            "MedianNotchDiff":  round(notch.median(), 2) if len(notch) else None,
            "MinNotchDiff":     notch.min()               if len(notch) else None,
            "MaxNotchDiff":     notch.max()               if len(notch) else None,
            "EntityIds":        ", ".join(entity_ids),
            "TopKeywords":      top_kw,
            "SampleReasons":    " | ".join(samples),
        })

    summary = pd.DataFrame(rows).sort_values(
        ["Count", "AvgNotchDiff"], ascending=[False, True]
    )
    return summary


def build_entity_rollup(detail: pd.DataFrame) -> pd.DataFrame:
    """Per-cluster breakdown by EntityId."""
    if detail.empty:
        return pd.DataFrame(columns=[
            "Category", "EntityId", "BorrowerName", "OverrideCount", "AvgNotchDiff",
        ])

    rollup = (
        detail.groupby(["Category", "EntityId"], dropna=False)
        .agg(
            BorrowerName=("BorrowerName", lambda x: " / ".join(
                pd.Series(x).dropna().astype(str).unique().tolist()[:3]
            )),
            OverrideCount=("EntityId", "size"),
            AvgNotchDiff=("NotchDifference", "mean"),
        )
        .reset_index()
    )
    rollup["AvgNotchDiff"] = rollup["AvgNotchDiff"].round(2)
    return rollup.sort_values(["Category", "OverrideCount"], ascending=[True, False])


# ===========================================================================
# Excel export with light formatting
# ===========================================================================

def export_excel(
    detail: pd.DataFrame,
    summary: pd.DataFrame,
    entity_rollup: pd.DataFrame,
    output_path: Path,
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        detail.sort_values(
            ["Category", "EntityId"], ascending=[True, True]
        ).to_excel(writer, sheet_name="Detailed Results", index=False)

        summary.to_excel(writer, sheet_name="Group Summary", index=False)
        entity_rollup.to_excel(writer, sheet_name="Entity Rollup", index=False)

        # Run info sheet
        pd.DataFrame({
            "Setting": [
                "Input file", "Sheet used", "Override-only rows",
                "N clusters", "Auto-select K", "n-gram range",
                "sklearn used",
            ],
            "Value": [
                config.INPUT_FILE, config.SHEET_NAME, config.ONLY_OVERRIDE_ROWS,
                config.N_CLUSTERS, config.AUTO_SELECT_K, str(config.NGRAM_RANGE),
                SKLEARN_AVAILABLE,
            ],
        }).to_excel(writer, sheet_name="Run Info", index=False)

    # Formatting pass
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(output_path)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        # Auto-width (capped)
        for col_cells in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in col_cells
            )
            ws.column_dimensions[
                get_column_letter(col_cells[0].column)
            ].width = min(max(max_len + 2, 12), 60)
        # Wrap text in SampleReasons and EntityIds columns
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and len(str(cell.value)) > 60:
                    cell.alignment = Alignment(wrap_text=True)

    wb.save(output_path)


# ===========================================================================
# Entry point
# ===========================================================================

def main() -> None:
    base = Path(__file__).resolve().parent
    input_path  = base / config.INPUT_FILE
    output_path = base / config.OUTPUT_FILE

    if not input_path.exists():
        sys.exit(
            f"ERROR: Input file not found: {input_path}\n"
            f"Place the Excel file in the same folder as this script, "
            f"or update INPUT_FILE in config.py."
        )

    if not SKLEARN_AVAILABLE:
        print(
            "INFO: scikit-learn is not installed — using built-in numpy implementation.\n"
            "      For better performance: pip install scikit-learn"
        )
    else:
        print("INFO: scikit-learn detected — using optimised TF-IDF/KMeans.")

    print(f"\nLoading: {input_path.name}")
    raw_df = load_sheet(input_path, config.SHEET_NAME)
    print(f"  {len(raw_df):,} rows read from workbook")

    detail       = prepare_dataframe(raw_df)
    summary      = build_summary(detail)
    entity_rollup = build_entity_rollup(detail)

    export_excel(detail, summary, entity_rollup, output_path)

    print(f"\n{'='*55}")
    print(f"  Rows analysed    : {len(detail):,}")
    print(f"  Groups found     : {detail['Category'].nunique() if not detail.empty else 0}")
    print(f"  Output saved to  : {output_path}")
    print(f"{'='*55}")


if __name__ == "__main__":
    main()
